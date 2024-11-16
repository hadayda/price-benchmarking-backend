import os
import shutil
import tempfile

import requests
from openpyxl.reader.excel import load_workbook


def download_file(url):
    tempdir = tempfile.mkdtemp()

    url_split = url.split('/')
    filename = url_split[-1] or url_split[-2]
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    file_path = os.path.join(tempdir, filename)
    with open(file_path, mode='wb') as f:
        with requests.get(url, stream=True) as r:
            shutil.copyfileobj(r.raw, f)
            return file_path


def parse_xlxs(file_path='', file_url=''):
    if not file_path:
        file_path = download_file(file_url)
    workbook = load_workbook(file_path)
    worksheet = workbook.active

    data = []
    headers = [cell.value for cell in next(worksheet.iter_rows(max_row=1))]
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not all(row):
            continue
        row_dict = {headers[i]: row[i] for i in range(len(headers))}
        data.append(row_dict)
    return data
